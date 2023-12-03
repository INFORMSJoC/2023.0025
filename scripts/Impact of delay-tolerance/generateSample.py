# -*- coding: utf-8 -*-
"""
Created on 2018/3/26 下午4:56
author: Tong Jia
email: cecilio.jia@gmail.com
software: PyCharm
"""
import xlrd
import random
import numpy as np
from six.moves import xrange
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def generateSimulationSamples(numPositions=10, numSamples=3000):
    """Generate numSamples matrixs in shape of (10, 2), the first column corresponds to first category patient and second
    column corresponds to second category.

    Parameters:
    ----------
    :param numPositions: int
        The rows number of per matrix sample.
    :param numSamples: int
        The total number of matrixs.
    :return:
    """
    samples = []
    for sample in xrange(numSamples):
        matrix = []
        for position in xrange(numPositions):
            vec = []
            # Generate first category by a uniform distribution U[0, theta], where θ is also assumed to be uniformly distributed in [3,4]
            theta = random.uniform(a=3, b=4)
            value1 = random.uniform(a=0, b=theta)
            vec.append(value1)
            # Generate second category by a normal distribution N[2, sigma^2], where standard deviation sigma is uniformly distributed in [0, 1/3]
            sigma = random.uniform(a=0, b=1 / 3)
            value2 = random.normalvariate(mu=2, sigma=sigma)
            vec.append(value2)

            matrix.append(vec)
        samples.append(matrix)
    samples = np.asarray(a=samples, dtype=float)
    return samples

def write_simulation_samples_into_excel(samples, filename, sheetname):
    """
    Parameters:
    ----------
    :param samples: ndarray
    :param filename: str
    :param sheetname: str
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname

    headerfont = Font(name="Calibri", bold=True, size=12)
    headeralignment = Alignment(horizontal="left", vertical="center")
    contentalignment = Alignment(horizontal="right", vertical="center")

    for sample_index in xrange(samples.shape[0]):
        header = ws.cell(row=11 * sample_index + 1, column=1, value="Sample-%d" % (sample_index + 1))
        header.font = headerfont
        header.alignment = headeralignment
        for position_index in xrange(samples.shape[1]):
            for type_index in xrange(samples.shape[2]):
                ws.cell(row=11 * sample_index + 1 + position_index + 1, column=type_index + 1, value=samples[sample_index][position_index][type_index]).alignment = contentalignment

    wb.save(filename=filename)

def read_simulation_samples_from_excel(filename, sheetname="Sheet1", I=10, J=2, numSamples=3000):
    wb = xlrd.open_workbook(filename=filename)
    ws = wb.sheet_by_name(sheet_name=sheetname)

    samples = []
    for sample_index in xrange(numSamples):
        matrix = []
        for position_index in xrange(I):
            vec = []
            for type_index in xrange(J):
                vec.append(
                    ws.cell_value(
                        rowx=11 * sample_index + 1 + position_index,
                        colx=type_index
                    )
                )
            matrix.append(vec)
        samples.append(matrix)
    samples = np.asarray(a=samples, dtype=float)
    return samples


if __name__ == '__main__':
    samples = generateSimulationSamples()
    print(samples.shape)
    # write_simulation_samples_into_excel(samples=samples, filename="..//..//data//Simulation Samples.xlsx", sheetname="Sheet1")
    samplesRead = read_simulation_samples_from_excel(
        filename="..//..//data//Simulation Samples.xlsx"
    )